"""
 * Author: Penut
 * Date: 2018/01/29
"""

import time
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def dog_login(driver):
    """
     * 前往登入頁面並等待使用者手動登入
    """
    driver.get("https://statementdog.com/users/sign_in")
    print("Waiting for log in")
    input()

def dog_crawler(driver, stock_id):
    """
     * 前往 stock_id 指定的股票頁面爬取資料
    """
    # 前往股票頁面
    driver.get("https://statementdog.com/analysis/tpe/" + str(stock_id))

    # 點擊三大財務報表
    wait = WebDriverWait(driver, 10)
    # 確定主選單已經顯示再繼續動作
    target = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "menu-title")))
    target[1].click()
    time.sleep(1) # 等待一秒

    # 切換至年報
    driver.execute_script("switchPlotType(1)")
    # 點擊總資產
    menus = driver.find_elements_by_tag_name("li")
    click_element(menus, "總資產")
    
    # 切換顯示日期，點選自訂 > 設定起始年度為2001
    # 點擊顯示年分
    driver.find_element_by_class_name("sheet-interval-current-option-text").click()
    # 點擊自訂
    click_element(menus, "自訂")
    # 選取起始年度
    startyear = driver.find_element_by_name("startyear")
    # 預設起始年度下方的第一個Option為2001
    startyear.find_element_by_tag_name("option").click()
    # Apply custom date
    driver.execute_script("customDate()")

    # 取得資料報表
    wait.until(EC.presence_of_all_elements_located((By.ID, "dataTable")))
    data = data_to_2dlist(driver.find_element_by_id("dataTable").text)
    item = data_to_2dlist(driver.find_element_by_id("itemTable").text)

    # 列印資料到主控台
    for row, title in enumerate(item):
        print(title[0], end='\t')
        for col in range(0, len(data[0])):
            print(data[row][col], end='\t')
        print()
    
    # 將資料寫進工作表
    data_to_xlsx(item, data, stock_id)

    # print(data)
    # print(item)

# 將項目與資料內容寫入工作表
def data_to_xlsx(item, data, filename):
    wb = openpyxl.Workbook() # 新建工作表
    st = wb.active           # 取得第一個資料表
    row_idx = 1; col_idx = 1 # 初始化寫入座標

    # 開始寫入資料表
    for row, title in enumerate(item):
        # 寫入項目名稱並設定靠左對齊
        c = st.cell(row=row_idx, column=col_idx, value=title[0])
        c.alignment = Alignment(horizontal='left')
        col_idx += 1
        for col in range(0, len(data[0])):
            # 寫入資料內容
            c = st.cell(row=row_idx, column=col_idx, value=data[row][col])
            # First Row 是年度，置中對齊
            if row_idx == 1: c.alignment = Alignment(horizontal='center')
            # 其餘的 Row 為數值資料，靠右對齊
            else: c.alignment = Alignment(horizontal='right')
            col_idx += 1
        col_idx = 1
        row_idx += 1
    
    # 將欄寬設定為 16
    for i in range(0, len(data[0]) + 1):
        st.column_dimensions[get_column_letter(i+1)].width = 16
    # 儲存檔案
    wb.save(str(filename) + ".xlsx")

# 從元素列表中點擊指定目標
def click_element(elems, target):
    for elem in elems:
        if elem.text == target:
            elem.click()
            break

# 將 Raw Data 轉為二維列表
def data_to_2dlist(data):
    data_list = data.split("\n")
    data_frame = []
    for item in data_list:
        data_frame.append(item.split(" "))
    return data_frame

def __main__():
    driver = webdriver.Firefox()
    dog_crawler(driver, 2330) # 台積電
    dog_crawler(driver, 6263) # 普萊德
    dog_crawler(driver, 2317) # 鴻海
    driver.close()

__main__()

"""
Statement Dog Function Reference

三大財務報表
gVar.analysis.stock.createPlotForMonthSales()    每月營收
gVar.analysis.stock.createPlotForEPS()           每股盈餘
gVar.analysis.stock.navPlot()                    每股淨值
gVar.analysis.stock.incomePlot()                 損益表
gVar.analysis.stock.createPlotForAssets()        總資產
gVar.analysis.stock.createPlotForDebtEquity()    負債和股東權益
gVar.analysis.stock.createPlotForCashflow()      現金流量表
gVar.analysis.stock.createPlotForDividend()      股利政策
gVar.analysis.stock.showEReport()                電子書

獲利能力分析
gVar.analysis.stock.createPlotForMargin()        利潤比率
gVar.analysis.stock.operatingExpenseRatioPlot()  營業費用率拆解
gVar.analysis.stock.nonOperatingIncomeRatePlot() 業外收支佔稅前淨利比例
gVar.analysis.stock.createPlotForReturnRatio()   ROE / ROA
gVar.analysis.stock.createPlotForDupont()        杜邦分析
gVar.analysis.stock.createPlotForTurnover()      經營週轉能力
gVar.analysis.stock.turnoverIntvlPlot()          營運週轉天數
gVar.analysis.stock.operatingCFToNetincomePlot() 營業現金流對淨利比
gVar.analysis.stock.cashDividentRatePlot()       現金股利發放率

安全性分析
gVar.analysis.stock.createPlotForCapitalStructPlot() 財務結構比率
gVar.analysis.stock.liabilityRatioPlot()             流速動比率
gVar.analysis.stock.interestCoveragePlot()           利息保障倍數
gVar.analysis.stock.cashflowLiabilityPlot()          現金流量分析
gVar.analysis.stock.operatingCFToNetincomePlot()     營業現金流對淨利比
gVar.analysis.stock.investmentRatioPlot()            盈餘再投資比率

成長力分析
gVar.analysis.stock.createPlotForMonthSalesGrowth()    月營收成長率
gVar.analysis.stock.createPlotForSalesGrowth()         營收成長率
gVar.analysis.stock.createPlotForGrossProfitGrowth()   毛利成長率
gVar.analysis.stock.createPlotForOperateIncomeGrowth() 營業利益成長率
gVar.analysis.stock.createPlotForNetIncomeGrowth()     稅後淨利成長率
gVar.analysis.stock.createPlotForEPSGrowth()           每股盈餘成長率

企業價值評估
gVar.analysis.stock.PERPlot()                          本益比評價
gVar.analysis.stock.PEBandPlot()                       本益比河流圖
gVar.analysis.stock.PBRPlot()                          股價淨值比評價
gVar.analysis.stock.PBBandPlot()                       股價淨值比河流圖
gVar.analysis.stock.cashDividentYieldPlot()            現金股利殖利率
gVar.analysis.stock.avgCashDividendYieldPlot()         平均現金股息殖利率
gVar.analysis.stock.createPlotForAvgCashDividendBand() 平均現金股息河流圖

經營階層觀察
gVar.analysis.stock.boardSharePlot()      董監持股比例
gVar.analysis.stock.mortgagePercentPlot() 董監持股質押比例
gVar.analysis.stock.boardBonusPlot()      董監酬金觀察
"""
